import openpyxl, random, string 

wb = openpyxl.load_workbook('file.xlsx')

ws = wb['namasheet']

def generate_random_password():
    
  '''
    Kriteria Password: 
    1. Panjang 8 karakter
    2. 5 karakter awal merupakan huruf, 3 sisanya angka
    3. huruf dengan index ganjil selalu huruf vocal, index genap consonant
    4. Semua huruf lowercase
  '''

  vowels = ['a', 'i', 'u', 'e', 'o'] 
  consonants = [i for i in string.ascii_lowercase if i not in vowels]

  word = ''.join(random.choice(vowels if i % 2 == 0 else consonants) for i in range(5))
  number = ''.join(random.choice(string.digits) for j in range(3))

  password = word + number
  return password


def insert_to_excel(start_row_pos, password):

  # set position to insert
  row_pos = start_row_pos
  column_pos = 4

  ws.cell(row_pos, column_pos).value = password


def main():

  # Dimulai dari row 2 karena row 1 adalah judul kolom
  for i in range(2, ws.max_row + 1):
    pw = generate_random_password()
    insert_to_excel(i, pw)
  
  wb.save('Akun pemira.xlsx')

  print("Proses Input Password Selesai")

main()


# print('The value in cell A1 is: '+ws['B2'].value)

# values = [ws.cell(row=i,column=2).value for i in range(2,7)]
# print(values)

# ws['D1'] = 'Password'

# wb.save('Gabungan.xlsx')